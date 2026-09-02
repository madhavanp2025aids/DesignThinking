"""
HYDAC Spec-to-3D Generator — XLSX Parser
Atomic, testable module: one file type only.
Uses openpyxl for merged cell handling + pandas for data extraction.
"""

from typing import Optional


def parse_xlsx(file_path: str, filename: str) -> dict:
    """
    Parse an XLSX file, extracting data from all sheets.
    Handles merged cells by distributing the top-left value to all cells in the range.

    Returns:
        {
            "raw_text": str,
            "tables": [{"table_index": int, "source_location": str, "headers": [...], "rows": [[...]]}],
            "sheets": [{"name": str, "tables": [...]}]
        }
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX parsing. Install: pip install openpyxl")

    wb = load_workbook(file_path, data_only=True)

    result = {
        "raw_text": "",
        "tables": [],
        "sheets": [],
    }

    table_index = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "name": sheet_name,
            "tables": [],
        }

        # Step 1: Unmerge cells and fill with top-left value
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = top_left_value

        # Step 2: Extract all data from the sheet
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert to strings, handle None
            str_row = [str(cell).strip() if cell is not None else "" for cell in row]
            # Skip completely empty rows
            if any(cell != "" for cell in str_row):
                all_rows.append(str_row)

        if not all_rows:
            continue

        # Step 3: Detect table boundaries (groups of consecutive non-empty rows)
        tables = _detect_tables(all_rows, sheet_name, table_index)

        for table in tables:
            sheet_data["tables"].append(table)
            result["tables"].append(table)
            table_index += 1

        # Raw text: concat all non-empty cells
        for row in all_rows:
            result["raw_text"] += " | ".join(cell for cell in row if cell) + "\n"

        result["sheets"].append(sheet_data)

    wb.close()
    return result


def _detect_tables(rows: list, sheet_name: str, start_index: int) -> list:
    """
    Detect table structures in a list of rows.
    Heuristic: first row with content is headers, subsequent rows are data,
    empty row or significant format change marks table boundary.
    """
    if not rows:
        return []

    tables = []
    current_headers = None
    current_rows = []
    header_col_count = 0

    for row in rows:
        non_empty_count = sum(1 for cell in row if cell)

        if current_headers is None:
            # This is a potential header row
            if non_empty_count >= 2:
                current_headers = row
                header_col_count = non_empty_count
                current_rows = []
        else:
            # Check if this row belongs to the current table
            if non_empty_count == 0:
                # Empty row = table boundary
                if current_rows:
                    tables.append({
                        "table_index": start_index + len(tables),
                        "source_location": f"Sheet '{sheet_name}'",
                        "headers": current_headers,
                        "rows": current_rows,
                    })
                current_headers = None
                current_rows = []
            else:
                current_rows.append(row)

    # Don't forget the last table
    if current_headers and current_rows:
        tables.append({
            "table_index": start_index + len(tables),
            "source_location": f"Sheet '{sheet_name}'",
            "headers": current_headers,
            "rows": current_rows,
        })

    # If no tables detected but we have rows, treat entire sheet as one table
    if not tables and len(rows) >= 2:
        tables.append({
            "table_index": start_index,
            "source_location": f"Sheet '{sheet_name}'",
            "headers": rows[0],
            "rows": rows[1:],
        })

    return tables
